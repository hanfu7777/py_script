import re
from openpyxl import load_workbook


# 我们使用的是openpyxl中的load_workbook模块来处理excel中的测试数据
# 这个是处理excel的数据的py文件

def get_test_data_from_excel(file, sheet_name) -> list:
    """
    :param file: excel文件名
    :param sheet_name: excel的sheet页
    :return:以列表的形式展示测试数据，每条测试数据就是一个字典
    """
    wb = load_workbook(filename=file)
    sh = wb[sheet_name]
    row = sh.max_row
    # 读取最大行
    column = sh.max_column
    # 读取最大列
    data = []
    keys = []
    for i in range(1, column + 1):
        keys.append(sh.cell(1, i).value)
        # 读取标题作为字典的key，此时keys列表均为标题
    for i in range(2, row + 1):
        temp = {}
        for j in range(1, column + 1):
            temp[keys[j - 1]] = sh.cell(i, j).value
        # temp['json'] = json.loads(temp['json'])
        # temp['expected'] = json.loads(temp['expected'])
        # temp['headers'] = json.loads(temp['headers'])
        data.append(temp)
    return data


def replace_args_by_re(json_s, obj):
    """
    :param json_s: 被查找的字符串,会动态查找##包裹的内容,并以列表的形式返回该内容
    :param obj:
    :return: 替换好的字符串
    """
    # 1、找出所有的槽位中的变量名
    args = re.findall('#(.+?)#', json_s)
    # print(args,type(args))
    # 2、再到obj中找到对应的属性替换
    for arg in args:
        # 遍历这个列表
        value = getattr(obj, arg, None)
        # 类.方法名利用反射查找这个属性 没有返回None
        # print(value)
        if value:
            json_s = json_s.replace('#{0}#'.format(arg), str(value))
            print('替换成功,新内容为{}'.format(str(value)))
    return json_s


if __name__ == '__main__':
    import os


    # base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # res = get_test_data_from_excel(os.path.join(base_dir, r'TestData\TestCases.xlsx'), 'Test')
    # print(res, type(res))
    # ..表示上级目录就是多层级安全存储
    # print(os.path.join(base_dir, "TestData\Testcases.xlsx"))

    class Env:
        name = 'han'
        age = 24
        bucketName = 1161805431


    s = '{"bucketName": "#bucketName#", "policy": "#name#", "server": "minio", "users": []}'
    res = replace_args_by_re(s, Env)
    print(res, type(res))
